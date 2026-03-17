#!/usr/bin/env python3
"""
Report Generator v1.0 - Tam Rapor Olusturma

8 Soru Raporu:
1. Ana konular nedir? (LLM)
2. Kac takipcisi var? (Profil)
3. Kac kisi takip ediyor? (Profil)
4. Takipci degisimi (Profil history)
5. Etkilesim degisimi (Metrics)
6. Parti/lider savunusu (LLM)
7. Muhalefet elestirisi (LLM)
8. En cok etkilesim alan tweetler (Metrics)
"""

from datetime import datetime
from typing import Dict, List, Optional

from app.core.database import (
    get_latest_profile,
    get_report_cache,
    save_report_cache,
    clear_report_cache
)
from app.core.constants import normalize_party_name
from app.services.scraping.profile_scraper import get_weekly_comparison
from app.services.reporting.metrics import (
    get_user_engagement_stats,
    get_top_tweets,
    compare_last_weeks
)


class ReportGenerator:
    """Tam rapor olusturucu"""

    def __init__(self, use_cache: bool = True, use_llm: bool = True):
        """
        Args:
            use_cache: Cache kullan (default True)
            use_llm: LLM analizi yap (default True, False = sadece metrikler)
        """
        self.use_cache = use_cache
        self.use_llm = use_llm
        self.analyzer = None

        if use_llm:
            try:
                print("[DEBUG] TweetAnalyzer import ediliyor...")
                from app.services.analysis.analyzer import TweetAnalyzer
                print("[DEBUG] TweetAnalyzer olusturuluyor...")
                self.analyzer = TweetAnalyzer()
                print(f"[DEBUG] TweetAnalyzer hazir, model: {self.analyzer.model}")
            except Exception as e:
                import traceback
                print(f"LLM baslatilamadi: {e}")
                print(f"[DEBUG] Traceback: {traceback.format_exc()}")
                self.use_llm = False

    def generate_report(self, username: str, force_refresh: bool = False) -> str:
        """
        Kullanici icin tam rapor olustur

        Args:
            username: Twitter kullanici adi
            force_refresh: Cache'i yoksay ve yeniden olustur

        Returns:
            Markdown formatinda rapor
        """
        # Cache kontrol
        if self.use_cache and not force_refresh:
            cached = get_report_cache(username, 'full')
            if cached:
                print(f"  Cache'den yuklendi (olusturulma: {cached['created_at']})")
                return cached['content']

        print(f"\n{'='*60}")
        print(f"RAPOR OLUSTURULUYOR: @{username}")
        print(f"{'='*60}")

        report_parts = []
        report_parts.append(self._generate_header(username))

        # Soru 2-3: Profil bilgileri
        print("  [1/6] Profil bilgileri...")
        report_parts.append(self._generate_profile_section(username))

        # Soru 4: Takipci degisimi
        print("  [2/6] Takipci degisimi...")
        report_parts.append(self._generate_follower_change_section(username))

        # Soru 5: Etkilesim degisimi
        print("  [3/6] Etkilesim metrikleri...")
        report_parts.append(self._generate_engagement_section(username))

        # Soru 8: En iyi tweetler
        print("  [4/6] En iyi tweetler...")
        report_parts.append(self._generate_top_tweets_section(username))

        # LLM Analizi (Soru 1, 6, 7)
        if self.use_llm and self.analyzer:
            print("  [5/6] LLM analizi (bu biraz zaman alabilir)...")
            report_parts.append(self._generate_llm_analysis_section(username))
        else:
            report_parts.append("\n## LLM Analizi\n*LLM analizi devre disi*\n")

        # Footer
        print("  [6/6] Rapor tamamlaniyor...")
        report_parts.append(self._generate_footer())

        # Birlestir
        report = "\n".join(report_parts)

        # Cache'e kaydet
        if self.use_cache:
            save_report_cache(username, 'full', report)
            print("  Rapor cache'e kaydedildi")

        print(f"{'='*60}")
        print("RAPOR TAMAMLANDI")
        print(f"{'='*60}\n")

        return report

    def _generate_header(self, username: str) -> str:
        """Rapor basligi"""
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        return f"""# Twitter Analiz Raporu: @{username}

**Olusturulma Tarihi:** {now}

---
"""

    def _generate_profile_section(self, username: str) -> str:
        """Soru 2-3: Profil bilgileri"""
        profile = get_latest_profile(username)

        if not profile:
            return """## Profil Bilgileri

*Profil verisi bulunamadi*
"""

        return f"""## Profil Bilgileri

| Metrik | Deger |
|--------|-------|
| Takipci | **{profile['followers']:,}** |
| Takip | **{profile['following']:,}** |
| Tweet | **{profile['tweets']:,}** |
| Son Guncelleme | {profile['date']} |
"""

    def _generate_follower_change_section(self, username: str) -> str:
        """Soru 4: Takipci degisimi"""
        comparison = get_weekly_comparison(username)

        if not comparison:
            return """## Takipci Degisimi

*Karsilastirma icin yeterli veri yok*
"""

        followers_change = comparison['followers_change']
        change_emoji = "📈" if followers_change > 0 else "📉" if followers_change < 0 else "➡️"
        change_str = f"+{followers_change:,}" if followers_change >= 0 else f"{followers_change:,}"

        return f"""## Takipci Degisimi (Haftalik)

| Metrik | Baslangic | Son | Degisim |
|--------|-----------|-----|---------|
| Takipci | {comparison['followers_start']:,} | {comparison['followers_end']:,} | {change_emoji} **{change_str}** |
| Takip | - | - | {comparison['following_change']:+,} |
| Tweet | - | - | {comparison['tweets_change']:+,} |

**Donem:** {comparison['period_start']} - {comparison['period_end']}
"""

    def _generate_engagement_section(self, username: str) -> str:
        """Soru 5: Etkilesim metrikleri"""
        stats = get_user_engagement_stats(username)
        weekly = compare_last_weeks(username, weeks=1)

        if stats['tweet_count'] == 0:
            return """## Etkilesim Metrikleri

*Tweet verisi bulunamadi*
"""

        # Haftalik degisim
        change_section = ""
        if weekly and weekly['period1']['total_engagement'] > 0:
            change_pct = weekly['changes']['engagement_change_pct']
            change_emoji = "📈" if change_pct > 0 else "📉" if change_pct < 0 else "➡️"
            change_section = f"\n**Haftalik Degisim:** {change_emoji} {change_pct:+.1f}%"

        return f"""## Etkilesim Metrikleri

| Metrik | Toplam | Tweet Basina |
|--------|--------|--------------|
| Like | {stats['total_likes']:,} | {stats['total_likes'] / max(stats['tweet_count'], 1):.1f} |
| Reply | {stats['total_replies']:,} | {stats['total_replies'] / max(stats['tweet_count'], 1):.1f} |
| Retweet | {stats['total_retweets']:,} | {stats['total_retweets'] / max(stats['tweet_count'], 1):.1f} |
| View | {stats['total_views']:,} | {stats['total_views'] / max(stats['tweet_count'], 1):.0f} |

**Toplam Tweet:** {stats['tweet_count']}
**Toplam Engagement:** {stats['total_engagement']:,}
**Ort. Engagement/Tweet:** {stats['avg_engagement_per_tweet']:.1f}
**Engagement Rate:** {stats['avg_engagement_rate']:.2f}%{change_section}
"""

    def _generate_top_tweets_section(self, username: str, limit: int = 5) -> str:
        """Soru 8: En iyi tweetler"""
        tweets = get_top_tweets(username, limit=limit)

        if not tweets:
            return """## En Cok Etkilesim Alan Tweetler

*Tweet bulunamadi*
"""

        lines = ["## En Cok Etkilesim Alan Tweetler\n"]

        for i, t in enumerate(tweets, 1):
            text = t['text']
            lines.append(f"""### {i}. Tweet
> {text}

- **Tarih:** {t['date'] or 'Bilinmiyor'}
- **Like:** {t['likes']:,} | **Reply:** {t['replies']:,} | **RT:** {t['retweets']:,}
- **Toplam Engagement:** {t['engagement']:,}
""")

        return "\n".join(lines)

    def _generate_llm_analysis_section(self, username: str) -> str:
        """Soru 1, 6, 7: LLM analizi (orijinal tweetler + retweetler)"""
        if not self.analyzer:
            return "\n## LLM Analizi\n*LLM kullanilabilir degil*\n"

        # Tweetleri ve parti bilgisini al
        import sqlite3
        from app.core.config import DB_PATH

        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # Parti bilgisini al ve normalize et
        cursor.execute("SELECT party FROM councilors WHERE username = ?", (username,))
        party_row = cursor.fetchone()
        party = normalize_party_name(party_row[0]) if party_row else "Bağımsız"

        # Orijinal tweetleri al - Metriklerle birlikte
        cursor.execute("""
            SELECT tweet_text, tweet_date, likes, views
            FROM tweets
            WHERE username = ? AND is_retweet = 0
            ORDER BY tweet_date DESC
        """, (username,))
        original_rows = cursor.fetchall()

        # Retweetleri al
        cursor.execute("""
            SELECT tweet_text, tweet_date, retweet_from
            FROM tweets
            WHERE username = ? AND is_retweet = 1
            ORDER BY tweet_date DESC
        """, (username,))
        retweet_rows = cursor.fetchall()
        conn.close()

        if not original_rows and not retweet_rows:
            return "\n## Profesyonel Istihbarat Analizi\n*Analiz icin tweet bulunamadi*\n"

        tweets = [
            {
                'tweet_text': row[0],
                'tweet_date': row[1],
                'likes': row[2] or 0,
                'views': row[3] or 0
            } for row in original_rows
        ]

        retweets = [
            {
                'tweet_text': row[0],
                'tweet_date': row[1],
                'retweet_from': row[2] or 'bilinmiyor'
            } for row in retweet_rows
        ]

        # Tarih araligi
        all_dates = [t['tweet_date'] for t in tweets if t['tweet_date']] + \
                    [t['tweet_date'] for t in retweets if t['tweet_date']]
        period = f"{min(all_dates)[:10]} - {max(all_dates)[:10]}" if all_dates else "Bilinmiyor"

        # Profesyonel İstihbarat Analizi Yap
        try:
            print(f"[DEBUG] analyze_intelligence cagiriliyor, {len(tweets)} tweet + {len(retweets)} RT, parti: {party}...")
            result = self.analyzer.analyze_intelligence(tweets, username, period=period, party=party, retweets=retweets)

            if not result.get('validated') or not result.get('analysis'):
                # LLM calismiyor veya JSON parse edilemedi
                error_msg = result.get('error', 'Bilinmeyen hata')
                return f"""
## Istihbarat Analizi

> LLM analizi yapilamadi. Hata: {error_msg[:200]}

**Parti:** {party}
**Orijinal Tweet Sayisi:** {len(tweets)}
**Retweet Sayisi:** {len(retweets)}
**Donem:** {period}

*LLM Provider: {self.analyzer.provider} | Model: {self.analyzer.model}*
"""

            analysis = result['analysis']

            # Markdown oluşturma
            report = []
            report.append("\n## Profesyonel Istihbarat Analizi")
            confidence = getattr(analysis, 'confidence_score', 0.7)
            confidence_label = "Yuksek" if confidence >= 0.8 else "Orta" if confidence >= 0.6 else "Dusuk"
            report.append(f"\n**Analiz Guveni:** {confidence_label} ({confidence:.0%})")
            report.append(f"\n**Kapsam:** {len(tweets)} orijinal tweet + {len(retweets)} retweet analiz edildi")
            report.append(f"\n> {analysis.executive_summary}")

            # 1. Yeşil Takım
            report.append("\n### Yesil Takim (Parti Temsili ve Sadakat)")
            report.append(f"**Sadakat Duzeyi:** {analysis.loyalty_level}")
            report.append(f"\n**Analiz:** {analysis.green_summary}")

            # 2. Kırmızı Takım
            report.append("\n### Kirmizi Takim (Rakip Analizi ve Elestiri)")
            report.append(f"**Elestiri Duzeyi:** {analysis.criticism_level}")
            report.append(f"\n**Analiz:** {analysis.red_summary}")

            # 3. Gri Takım
            report.append("\n### Gri Takim (Bagimsiz Alan)")
            report.append(f"\n**Analiz:** {analysis.grey_summary}")

            if analysis.independent_topics:
                report.append("\n**Bagimsiz Gundemler:**")
                for topic in analysis.independent_topics:
                    report.append(f"- {topic}")

            # 4. Retweet Analizi
            retweet_summary = getattr(analysis, 'retweet_summary', '')
            retweet_sources = getattr(analysis, 'retweet_sources', [])

            if retweet_summary or retweet_sources:
                report.append("\n### Retweet Analizi")
                report.append(f"**RT Sayisi:** {len(retweets)} tweet")
                if retweet_summary:
                    report.append(f"\n**Analiz:** {retweet_summary}")
                if retweet_sources:
                    report.append("\n**Sik RT Edilen Kaynaklar:**")
                    for source in retweet_sources[:10]:
                        report.append(f"- @{source}")

            return "\n".join(report) + "\n"

        except Exception as e:
            return f"\n## Istihbarat Analizi\n\n*LLM analizi yapilamadi. Hata: {str(e)[:100]}*\n"

    def _generate_footer(self) -> str:
        """Rapor footer"""
        return """
---

*Bu rapor Meclis Istihbarat Sistemi tarafindan otomatik olusturulmustur.*
*Rapor suresi: 1 hafta gecerlidir.*
"""


# ============================================================================
# YARDIMCI FONKSIYONLAR
# ============================================================================

def generate_report(username: str, use_cache: bool = True, use_llm: bool = True) -> str:
    """Kisa yol: Rapor olustur"""
    generator = ReportGenerator(use_cache=use_cache, use_llm=use_llm)
    return generator.generate_report(username)


def generate_reports_batch(usernames: List[str], use_llm: bool = True) -> Dict[str, str]:
    """Birden fazla kullanici icin rapor olustur"""
    generator = ReportGenerator(use_cache=True, use_llm=use_llm)
    reports = {}

    for i, username in enumerate(usernames, 1):
        print(f"\n[{i}/{len(usernames)}] @{username}")
        try:
            reports[username] = generator.generate_report(username)
        except Exception as e:
            print(f"  Hata: {e}")
            reports[username] = f"# Hata\n\nRapor olusturulamadi: {e}"

    return reports


def generate_quick_report(username: str) -> str:
    """LLM olmadan hizli rapor"""
    generator = ReportGenerator(use_cache=True, use_llm=False)
    return generator.generate_report(username)


# ============================================================================
# EXPORT FONKSIYONLARI
# ============================================================================

def export_to_excel(data: List[Dict], filename: str) -> Optional[str]:
    """
    Veriyi Excel dosyasina aktar

    Args:
        data: Liste halinde sozluk verileri
        filename: Cikti dosya adi (.xlsx uzantisi olmadan)

    Returns:
        Olusturulan dosya yolu
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows

        df = pd.DataFrame(data)
        filepath = f"{filename}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Rapor"

        # Header stilleri
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # DataFrame'i Excel'e yaz
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

        # Sutun genislikleri
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        wb.save(filepath)
        return filepath

    except Exception as e:
        print(f"Excel export hatasi: {e}")
        return None


def export_to_pdf(markdown_content: str, filename: str, title: str = "Rapor") -> Optional[str]:
    """
    Markdown icerigini PDF'e aktar

    Args:
        markdown_content: Markdown formatinda metin
        filename: Cikti dosya adi (.pdf uzantisi olmadan)
        title: PDF basligi

    Returns:
        Olusturulan dosya yolu
    """
    try:
        from fpdf import FPDF
        import re

        filepath = f"{filename}.pdf"

        pdf = FPDF()
        pdf.add_page()

        # Font ayarlari - Turkce karakter destegi icin
        pdf.add_font('DejaVu', '', 'C:/Windows/Fonts/arial.ttf', uni=True)
        pdf.set_font('DejaVu', '', 12)

        # Baslik
        pdf.set_font('DejaVu', '', 18)
        pdf.cell(0, 15, title, ln=True, align='C')
        pdf.ln(5)

        # Markdown'i satirlara bol
        lines = markdown_content.split('\n')

        for line in lines:
            line = line.strip()

            if not line:
                pdf.ln(3)
                continue

            # Baslik kontrolu
            if line.startswith('# '):
                pdf.set_font('DejaVu', '', 16)
                pdf.multi_cell(0, 8, line[2:])
                pdf.ln(2)
            elif line.startswith('## '):
                pdf.set_font('DejaVu', '', 14)
                pdf.multi_cell(0, 7, line[3:])
                pdf.ln(2)
            elif line.startswith('### '):
                pdf.set_font('DejaVu', '', 12)
                pdf.multi_cell(0, 6, line[4:])
                pdf.ln(1)
            elif line.startswith('---'):
                pdf.ln(3)
                pdf.line(10, pdf.get_y(), 200, pdf.get_y())
                pdf.ln(3)
            elif line.startswith('> '):
                pdf.set_font('DejaVu', '', 10)
                pdf.set_x(15)
                pdf.multi_cell(0, 5, line[2:])
                pdf.set_x(10)
            elif line.startswith('|'):
                # Tablo satiri - basitce yaz
                clean_line = line.replace('|', ' ').strip()
                if clean_line and not all(c in '-| ' for c in line):
                    pdf.set_font('DejaVu', '', 10)
                    pdf.multi_cell(0, 5, clean_line)
            elif line.startswith('- ') or line.startswith('* '):
                pdf.set_font('DejaVu', '', 11)
                pdf.multi_cell(0, 5, "  " + line)
            else:
                # Normal metin
                pdf.set_font('DejaVu', '', 11)
                # Bold isaretlerini kaldir
                clean = re.sub(r'\*\*(.*?)\*\*', r'\1', line)
                pdf.multi_cell(0, 5, clean)

        pdf.output(filepath)
        return filepath

    except Exception as e:
        print(f"PDF export hatasi: {e}")
        return None


def export_engagement_excel(usernames: Optional[List[str]] = None) -> Optional[str]:
    """
    Tum kullanicilarin engagement verilerini Excel'e aktar

    Args:
        usernames: Kullanici listesi (None = hepsi)

    Returns:
        Dosya yolu
    """
    import sqlite3
    from app.core.config import DB_PATH
    from datetime import datetime

    conn = sqlite3.connect(DB_PATH)

    if usernames:
        placeholders = ','.join(['?' for _ in usernames])
        query = f"""
            SELECT
                c.username,
                c.name,
                c.party,
                c.district,
                COALESCE(ph.followers_count, 0) as followers,
                COUNT(t.id) as tweet_count,
                COALESCE(SUM(t.likes), 0) as total_likes,
                COALESCE(SUM(t.retweets), 0) as total_retweets,
                COALESCE(SUM(t.replies), 0) as total_replies,
                COALESCE(SUM(t.views), 0) as total_views
            FROM councilors c
            LEFT JOIN profile_history ph ON c.username = ph.username
            LEFT JOIN tweets t ON c.username = t.username AND t.is_retweet = 0
            WHERE c.username IN ({placeholders})
            GROUP BY c.username
            ORDER BY total_likes DESC
        """
        cursor = conn.execute(query, usernames)
    else:
        query = """
            SELECT
                c.username,
                c.name,
                c.party,
                c.district,
                COALESCE(ph.followers_count, 0) as followers,
                COUNT(t.id) as tweet_count,
                COALESCE(SUM(t.likes), 0) as total_likes,
                COALESCE(SUM(t.retweets), 0) as total_retweets,
                COALESCE(SUM(t.replies), 0) as total_replies,
                COALESCE(SUM(t.views), 0) as total_views
            FROM councilors c
            LEFT JOIN profile_history ph ON c.username = ph.username
            LEFT JOIN tweets t ON c.username = t.username AND t.is_retweet = 0
            GROUP BY c.username
            ORDER BY total_likes DESC
        """
        cursor = conn.execute(query)

    rows = cursor.fetchall()
    conn.close()

    data = [
        {
            'Username': row[0],
            'Isim': row[1],
            'Parti': normalize_party_name(row[2]),
            'Ilce': row[3],
            'Takipci': row[4],
            'Tweet Sayisi': row[5],
            'Toplam Like': row[6],
            'Toplam RT': row[7],
            'Toplam Reply': row[8],
            'Toplam View': row[9],
            'Engagement': row[6] + row[7] + row[8]
        }
        for row in rows
    ]

    filename = f"meclis_rapor_{datetime.now().strftime('%Y%m%d_%H%M')}"
    return export_to_excel(data, filename)


# ============================================================================
# CLI
# ============================================================================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Rapor Olusturma")
    parser.add_argument("--user", help="Kullanici adi")
    parser.add_argument("--users", nargs="+", help="Birden fazla kullanici")
    parser.add_argument("--no-cache", action="store_true", help="Cache kullanma")
    parser.add_argument("--no-llm", action="store_true", help="LLM kullanma (hizli mod)")
    parser.add_argument("--clear-cache", action="store_true", help="Cache temizle")
    parser.add_argument("--output", help="Cikti dosyasi (.md)")

    args = parser.parse_args()

    if args.clear_cache:
        from app.core.database import clear_report_cache
        deleted = clear_report_cache()
        print(f"Cache temizlendi: {deleted} kayit silindi")

    elif args.user:
        report = generate_report(
            args.user,
            use_cache=not args.no_cache,
            use_llm=not args.no_llm
        )

        if args.output:
            with open(args.output, 'w', encoding='utf-8') as f:
                f.write(report)
            print(f"Rapor kaydedildi: {args.output}")
        else:
            print(report)

    elif args.users:
        reports = generate_reports_batch(args.users, use_llm=not args.no_llm)
        for username, report in reports.items():
            if args.output:
                filename = f"{args.output}_{username}.md"
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write(report)
                print(f"Kaydedildi: {filename}")
            else:
                print(report)
                print("\n" + "="*80 + "\n")

    else:
        print("Kullanim:")
        print("  python report_generator.py --user username")
        print("  python report_generator.py --user username --no-llm")
        print("  python report_generator.py --user username --output rapor.md")
        print("  python report_generator.py --users user1 user2 user3")
        print("  python report_generator.py --clear-cache")